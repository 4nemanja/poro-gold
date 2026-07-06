"""
Marketplace OS data ingestion.

Parses the real sales spreadsheets under source-docs/ into normalized JSON
under data/ingested/, and (if Supabase credentials are set as env vars)
pushes the same data into Supabase.

Re-runnable: run again any time new weeks of data are added to
PORO Tracker.xlsm. Never reads the "Credentials" sheet - that sheet is
skipped by name before any row is touched.

Usage:
    python scripts/ingest.py                # parse + write local JSON only
    python scripts/ingest.py --push         # also push to Supabase (requires env vars)

Env vars for --push:
    SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY
"""

import argparse
import json
import os
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "source-docs"
OUT_DIR = ROOT / "data" / "ingested"

PLATFORM_NAMES = ["GameBoost", "iGV", "Z2U", "G2G", "Playhub", "PlayerOK", "KupujemProdajem"]
PLATFORM_LOOKUP = {p.upper(): p for p in PLATFORM_NAMES}

WEEK_PDF_ROW_RE = re.compile(
    r"^(GameBoost|iGV|Z2U|G2G|Playhub|PlayerOK|KupujemProdajem)\s+"
    r"(\d+)\s+\$([\d.]+)\s+\$([\d.]+)\s+(\d+)\s+(\d+)(?:\s+\$([\d.]+))?\s*$"
)
WEEK_PDF_TOTAL_RE = re.compile(r"^Total\s+(\d+)\s+(\d+)\s+\$([\d.]+)\s*$")


def warn(msg: str):
    print(f"WARNING: {msg}", file=sys.stderr)


def normalize_order_id(raw: str) -> str:
    return re.sub(r"\s*-\s*", "-", raw.strip())


def parse_date(raw: str) -> str:
    """DD.MM.YYYY -> YYYY-MM-DD"""
    return datetime.strptime(raw.strip(), "%d.%m.%Y").strftime("%Y-%m-%d")


def normalize_platform(raw: str) -> str:
    return PLATFORM_LOOKUP.get(raw.strip().upper(), raw.strip())


def find_header_row(ws, marker_col_idx: int, marker_value: str, max_scan=60):
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row)):
        cell = row[marker_col_idx]
        if cell.value == marker_value:
            return cell.row
    return None


def read_sales_log(wb):
    """Parses the Sales Log sheet - the source of truth for individual orders."""
    ws = wb["Sales Log"]
    header_row = find_header_row(ws, 1, "Order ID")
    if header_row is None:
        warn("Sales Log: could not find header row (looked for 'Order ID' in column B)")
        return [], None

    date_range = None
    for row in ws.iter_rows(min_row=1, max_row=header_row):
        for cell in row:
            if isinstance(cell.value, str) and "Date Range:" in cell.value:
                m = re.search(r"Date Range:\s*([\d.]+)\s*-\s*([\d.]+)", cell.value)
                if m:
                    date_range = {"start": parse_date(m.group(1)), "end": parse_date(m.group(2))}

    orders = []
    quality_issues = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        order_id_raw = row[1].value
        if order_id_raw is None or str(order_id_raw).strip() == "":
            continue
        if str(order_id_raw).strip().upper() == "ORDER ID":
            continue  # a repeated header block further down the sheet

        order_id = normalize_order_id(str(order_id_raw))
        if order_id != str(order_id_raw).strip():
            quality_issues.append(f"order_id '{order_id_raw}' normalized to '{order_id}'")

        try:
            date = parse_date(str(row[2].value)) if row[2].value else None
        except ValueError:
            quality_issues.append(f"order {order_id}: unparseable date '{row[2].value}'")
            date = None

        cost = row[7].value
        sold_for = row[8].value
        profit = row[9].value
        if profit is None:
            quality_issues.append(f"order {order_id}: PROFIT AFTER FEES is blank in the sheet")

        supplier_paid_raw = row[11].value
        supplier_paid = None
        if isinstance(supplier_paid_raw, str) and supplier_paid_raw.strip():
            supplier_paid = supplier_paid_raw.strip().upper() in ("YES", "TRUE", "PAID", "Y")

        orders.append({
            "order_id": order_id,
            "date": date,
            "method": row[3].value,
            "platform": normalize_platform(str(row[4].value)) if row[4].value else None,
            "product": row[5].value,
            "supplier": row[6].value,
            "cost": float(cost) if isinstance(cost, (int, float)) else None,
            "sold_for": float(sold_for) if isinstance(sold_for, (int, float)) else None,
            "profit": float(profit) if isinstance(profit, (int, float)) else None,
            "status": row[10].value,
            "supplier_paid": supplier_paid,
            "notes": row[12].value,
            "source_row": row[1].row,
        })

    return orders, date_range, quality_issues


def parse_fee_pct(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw) * 100 if raw < 1 else float(raw)
    m = re.search(r"([\d.]+)\s*%", str(raw))
    return float(m.group(1)) if m else None


def read_platforms(wb):
    ws = wb["Platforms"]
    header_row = find_header_row(ws, 1, "Platforms")
    if header_row is None:
        warn("Platforms: could not find header row")
        return []
    platforms = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        name_raw = row[1].value
        if not name_raw:
            continue
        platforms.append({
            "name": normalize_platform(str(name_raw)),
            "offers": row[2].value,
            "active": str(row[3].value).strip().lower() == "active" if row[3].value else None,
            "description": row[4].value,
            "withdrawal_fee_pct": parse_fee_pct(row[5].value),
            "withdrawal_fee_raw": row[5].value,
            "withdrawal_method": row[6].value,
            "sync_status": "manual_import",
        })
    return platforms


def read_prices(wb):
    ws = wb["Prices"]
    header_row = find_header_row(ws, 1, "GAME")
    if header_row is None:
        warn("Prices: could not find header row")
        return []
    prices = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        game = row[1].value
        package = row[2].value
        if not game or not package:
            continue
        price = row[3].value
        prices.append({
            "game": game,
            "package": package,
            "price": float(price) if isinstance(price, (int, float)) else None,
            "profit_note": row[4].value,
            "cannot_sell_on": row[5].value,
        })
    return prices


def parse_week_pdf(path: Path, week_label: str):
    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    rows = []
    total_line = None
    for line in lines:
        m = WEEK_PDF_ROW_RE.match(line)
        if m:
            platform, sold, cost, revenue, refund, disputes, income = m.groups()
            rows.append({
                "week_label": week_label,
                "platform": platform,
                "total_sold": int(sold),
                "total_cost": float(cost),
                "total_sales_revenue": float(revenue),
                "total_refund": int(refund),
                "disputes": int(disputes),
                "total_income_after_fees": float(income) if income else 0.0,
            })
            continue
        m = WEEK_PDF_TOTAL_RE.match(line)
        if m:
            total_line = {
                "total_refund": int(m.group(1)),
                "disputes": int(m.group(2)),
                "total_income_after_fees": float(m.group(3)),
            }

    return rows, total_line


def cross_check_week(rows, total_line, week_label, issues):
    if total_line is None:
        issues.append(f"{week_label}: no 'Total' rollup line found in PDF to cross-check against")
        return
    summed_refund = sum(r["total_refund"] for r in rows)
    summed_disputes = sum(r["disputes"] for r in rows)
    summed_income = round(sum(r["total_income_after_fees"] for r in rows), 2)

    if summed_refund != total_line["total_refund"]:
        issues.append(
            f"{week_label}: MISMATCH refund total - platforms sum to {summed_refund}, "
            f"PDF Total row says {total_line['total_refund']}"
        )
    if summed_disputes != total_line["disputes"]:
        issues.append(
            f"{week_label}: MISMATCH disputes total - platforms sum to {summed_disputes}, "
            f"PDF Total row says {total_line['disputes']}"
        )
    if abs(summed_income - total_line["total_income_after_fees"]) > 0.01:
        issues.append(
            f"{week_label}: MISMATCH income-after-fees total - platforms sum to "
            f"${summed_income}, PDF Total row says ${total_line['total_income_after_fees']}"
        )


def push_to_supabase(table: str, rows: list, on_conflict: str, supabase_url: str, service_key: str):
    if not rows:
        return
    url = f"{supabase_url.rstrip('/')}/rest/v1/{table}?on_conflict={on_conflict}"
    body = json.dumps(rows).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("apikey", service_key)
    req.add_header("Authorization", f"Bearer {service_key}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "resolution=merge-duplicates,return=representation")
    try:
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read().decode("utf-8"))
            print(f"  pushed {len(result)} rows to '{table}'")
    except urllib.error.HTTPError as e:
        warn(f"push to '{table}' failed: {e.code} {e.read().decode('utf-8')}")
        raise


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--push", action="store_true", help="Push ingested data to Supabase")
    args = parser.parse_args()

    quality_issues = []

    # --- Excel sources ---
    tracker_path = SOURCE_DIR / "PORO Tracker.xlsm"
    if not tracker_path.exists():
        print(f"ERROR: {tracker_path} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(tracker_path, data_only=True, keep_vba=True)

    if "Credentials" in wb.sheetnames:
        warn("Credentials sheet found in PORO Tracker.xlsm - SKIPPING. Never read, never exported.")
    for name in wb.sheetnames:
        if name.lower() == "credentials":
            continue  # hard exclusion, enforced again defensively

    orders, current_week_range, order_issues = read_sales_log(wb)
    quality_issues.extend(order_issues)
    platforms = read_platforms(wb)
    prices = read_prices(wb)

    # --- PDF rollups (Week 1 / Week 2 - no order-level data exists for these) ---
    weekly_totals = []
    for label, filename in [("WEEK1", "WEEK1.pdf"), ("WEEK2", "WEEK2.pdf")]:
        path = SOURCE_DIR / filename
        if not path.exists():
            warn(f"{filename} not found, skipping")
            continue
        rows, total_line = parse_week_pdf(path, label)
        cross_check_week(rows, total_line, label, quality_issues)
        weekly_totals.extend(rows)

    # --- Suppliers derived from Sales Log ---
    supplier_names = sorted({o["supplier"] for o in orders if o["supplier"]})
    suppliers = [{"name": s, "status": "active"} for s in supplier_names]

    missing_profit = sum(1 for o in orders if o["profit"] is None)

    report = {
        "ingested_at": datetime.now(timezone.utc).isoformat(),
        "orders_count": len(orders),
        "current_week_range": current_week_range,
        "weekly_totals_rows": len(weekly_totals),
        "orders_missing_profit": missing_profit,
        "quality_issues": quality_issues,
    }

    # --- write local JSON (always, so the UI can run without Supabase) ---
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "fact_orders.json").write_text(json.dumps(orders, indent=2))
    (OUT_DIR / "dim_platform.json").write_text(json.dumps(platforms, indent=2))
    (OUT_DIR / "dim_prices.json").write_text(json.dumps(prices, indent=2))
    (OUT_DIR / "dim_suppliers.json").write_text(json.dumps(suppliers, indent=2))
    (OUT_DIR / "fact_weekly_platform_totals.json").write_text(json.dumps(weekly_totals, indent=2))
    (OUT_DIR / "ingestion_report.json").write_text(json.dumps(report, indent=2))

    print(f"Ingested {len(orders)} real orders ({current_week_range}), "
          f"{len(weekly_totals)} historical weekly-platform rollup rows, "
          f"{len(suppliers)} suppliers, {len(platforms)} platforms.")
    print(f"Wrote JSON to {OUT_DIR}")
    if quality_issues:
        print("\nData quality issues:")
        for issue in quality_issues:
            print(f"  - {issue}")

    if args.push:
        supabase_url = os.environ.get("SUPABASE_URL")
        service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        if not supabase_url or not service_key:
            print("\n--push given but SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set. Skipping push.")
            sys.exit(1)
        print("\nPushing to Supabase...")
        push_to_supabase("dim_platform", platforms, "name", supabase_url, service_key)
        push_to_supabase("dim_suppliers", suppliers, "name", supabase_url, service_key)
        push_to_supabase("fact_orders", orders, "order_id", supabase_url, service_key)
        push_to_supabase("fact_weekly_platform_totals", weekly_totals, "week_label,platform", supabase_url, service_key)
        print("Push complete. Query the tables back to confirm before trusting this.")
    else:
        print("\nRun with --push (and SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY set) to write to Supabase.")


if __name__ == "__main__":
    main()
